# Importing required libraries
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
import pandas as pd
from datetime import datetime
import os
from cloudfile_tech import delete_existing_files, upload_file_by_command


def writexlsxfile(hottokens):
    delete_existing_files('eigenphi_Hot_tokens.xlsx')

    wb = Workbook()
    sheet = wb.active
    sheet.title = "eth_bsc_hot_tokens"
    for i in range(2):
        if i == 0:
            sheet.merge_cells('A1:B1')
            merge_cell = sheet.cell(row=1, column=1)
            merge_cell.value = "Ethereum HOT Tokens"
        else:
            sheet.merge_cells('A35:B35')
            merge_cell = sheet.cell(row=35, column=1)
            merge_cell.value = "Binance HOT Tokens"

        merge_cell.alignment = Alignment(wrap_text=True, horizontal='center', vertical='center')
        merge_cell.font = Font(bold=True, italic=True)

        start_row = 2 if i == 0 else 37
        start_col = 1

        header_row = hottokens.columns.tolist()
        for j, col_name in enumerate(header_row):
            cell = sheet.cell(row=start_row, column=start_col + j)
            cell.value = col_name
            cell.font = Font(italic=True)
            cell.alignment = Alignment(horizontal='center', vertical='center')

        data_rows = hottokens[:30].values.tolist() if i == 0 else hottokens[30:].values.tolist()

        for k, row in enumerate(data_rows):
            for l, val in enumerate(row):
                cell = sheet.cell(row=start_row + k + 1, column=start_col + l)
                cell.value = val
                cell.alignment = Alignment(horizontal='left', vertical='center')

    columns = ['A', 'B']
    for col in columns:
        sheet.column_dimensions[col].width = 50 if col == 'A' else 18

    wb.save('eigenphi_Hot_tokens.xlsx')
    # get current working directory
    current_path = os.getcwd()

    upload_file_by_command(current_path, 'eigenphi_Hot_tokens.xlsx')

