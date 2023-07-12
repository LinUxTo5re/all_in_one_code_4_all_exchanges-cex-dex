"""
    This file will delete existing Excel file and
    write data to that file, also read data from
    that file.
"""

# Importing required libraried
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Alignment, Font
import pandas as pd
from datetime import datetime
import os


# delete file if already exists
def deletefileifexist(filename):
    if os.path.exists(filename):
        os.remove(filename)
        print(f'Information: {filename} exist. deletion started......')
        print(f"Information: {filename} has been deleted.")
    else:
        print(f"Information: {filename} does not exist.")


'''
    merege cell, write dataframe to excel file
    apply alignment, styles and many more
'''


def writexlsxfile(Compared_Markets_Data, filename, exchanges=None):
    # if file exist then first delete that file. file should not be opened
    # in any application. take care of that.
    deletefileifexist(filename)

    try:
        if exchanges is None:  # writing markets from multicex python file
            # Compared_Markets_Data.to_excel(filename, sheet_name='markets_from_multicex')
            # create a new workbook and select the active sheet
            wb = Workbook()
            sheet = wb.active
            # Write column headers to worksheet
            header_row = Compared_Markets_Data.columns.tolist()
            header_row.insert(0, 'Exchanges')
            for i, col_name in enumerate(header_row):
                cell = sheet.cell(row=1, column=1 + i)
                cell.value = col_name
                cell.font = Font(italic=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # concatinating values of exchanges
            exchanges_ = Compared_Markets_Data.index.to_list()
            data_rows = Compared_Markets_Data.values.tolist()
            for index, d in enumerate(data_rows):
                data_rows[index].insert(0, exchanges_[index])

            # write data rows to worksheet
            for i, row in enumerate(data_rows):
                for j, val in enumerate(row):
                    cell = sheet.cell(row=1 + i + 1, column=1 + j)
                    cell.value = val
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # sheet.column_dimensions['A':'Z'].width = 27
            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
            for col in columns:
                sheet.column_dimensions[col].width = 27
            wb.save(filename)

        else:  # writing markets from load_common_markets python file
            # create a new workbook and select the active sheet
            wb = Workbook()
            sheet = wb.active
            # current datetime
            current_datetime = datetime.now()
            current_datetime = current_datetime.strftime("%Y:%m:%d %H:%M")

            # merge cells A1:F8
            sheet.merge_cells('A1:F8')

            # add name to sheet
            sheet.title = 'Common_markets_bt_cex'

            # set the value of the merged cell to the paragraph text
            header_cell = sheet.cell(row=1, column=1)
            header_cell.value = f'''        first market: {exchanges[0]}
            second market: {exchanges[1]}
            datetime: {current_datetime}
            '''

            header_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
            header_cell.font = Font(bold=True, italic=True)

            # set the starting cell for data rows
            start_row = 9
            start_col = 1

            # add a header row
            header_row = Compared_Markets_Data.columns.tolist()
            for i, col_name in enumerate(header_row):
                cell = sheet.cell(row=start_row, column=start_col + i)
                cell.value = col_name
                cell.font = Font(italic=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # add data rows
            data_rows = Compared_Markets_Data.values.tolist()

            for i, row in enumerate(data_rows):
                for j, val in enumerate(row):
                    cell = sheet.cell(row=start_row + i + 1, column=start_col + j)
                    if isinstance(val, list):
                        cell.value = val[0]
                    else:
                        cell.value = val
                    cell.alignment = Alignment(horizontal='center', vertical='center')

            # sheet.column_dimensions['A':'Z'].width = 27
            columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
            for col in columns:
                sheet.column_dimensions[col].width = 27

            # save the workbook
            wb.save(filename)
    except Exception as e:
        print('Error: ' + str(e))
    else:
        print(f'{filename} written successfully')


# read existing Excel file from 8th row
def readxlsxfile(filename='common_market.xlsx'):
    if os.path.exists(filename):
        # Read the Excel file, skipping the first 5 rows
        df = pd.read_excel(filename, skiprows=8, index_col=None)
        # Print the DataFrame
        print(f'File: {filename} loaded successfully')
        return df
    else:
        print(f'Warning: {filename} not found')


# Entry point of every python file(optional)
if __name__ == '__main__':
    readxlsxfile('common_market.xlsx')
